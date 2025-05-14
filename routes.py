import calendar
from datetime import datetime, timedelta
from flask import render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from flask import render_template, redirect, url_for, flash, request
from models import ITInventory
from forms import ITInventoryForm
from app import db
from werkzeug.security import check_password_hash, generate_password_hash
from app import app, db
from models import User, SolderingTip, MachineCalibration, OvertimeLogbook, EquipmentDowntime
from forms import LoginForm, UserForm, SolderingTipForm, MachineCalibrationForm, OvertimeLogbookForm, EquipmentDowntimeForm, ReportForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from models import User, SolderingTip, MachineCalibration, OvertimeLogbook, EquipmentDowntime, MaintenanceReport
from forms import LoginForm, UserForm, SolderingTipForm, MachineCalibrationForm, OvertimeLogbookForm, EquipmentDowntimeForm, ReportForm, MaintenanceReportForm
from werkzeug.utils import secure_filename
import os
import json
# Home/Dashboard route
@app.route('/')
@login_required
def dashboard():
    
    # Count records for each module
    soldering_tips_count = SolderingTip.query.count()
    machine_calibrations_count = MachineCalibration.query.count()
    
    overtime_logs_count = OvertimeLogbook.query.count()
    equipment_downtimes_count = EquipmentDowntime.query.count()
    
    # Get recent records
    recent_soldering = SolderingTip.query.order_by(SolderingTip.created_at.desc()).limit(5).all()
    recent_calibrations = MachineCalibration.query.order_by(MachineCalibration.created_at.desc()).limit(5).all()
    recent_overtime = OvertimeLogbook.query.order_by(OvertimeLogbook.created_at.desc()).limit(5).all()
    recent_downtimes = EquipmentDowntime.query.order_by(EquipmentDowntime.created_at.desc()).limit(5).all()
    
    # Get upcoming calibrations
    upcoming_calibrations = MachineCalibration.query.filter(
        MachineCalibration.date.isnot(None),
        MachineCalibration.date <= datetime.now() + timedelta(days=7)
    ).all()
    
    return render_template('dashboard.html', 
                           soldering_tips_count=soldering_tips_count,
                           machine_calibrations_count=machine_calibrations_count,
                           overtime_logs_count=overtime_logs_count,
                           equipment_downtimes_count=equipment_downtimes_count,
                           recent_soldering=recent_soldering,
                           recent_calibrations=recent_calibrations,
                           recent_overtime=recent_overtime,
                           recent_downtimes=recent_downtimes,
                           upcoming_calibrations=upcoming_calibrations)

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password_hash, form.password.data):
            login_user(user, remember=form.remember_me.data)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# User management routes (admin only)
@app.route('/users')
@login_required
def user_management():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    search_query = request.args.get('search', '')
    if search_query:
        users = User.query.filter(
            User.username.ilike(f'%{search_query}%')
        ).all()
    else:
        users = User.query.all()
    form = UserForm()
    return render_template('user_management.html', users=users, form=form, search_query=search_query)

@app.route('/users/add', methods=['POST'])
@login_required
def add_user():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    form = UserForm()
    if form.validate_on_submit():
        # Check if username or email already exists
        if User.query.filter_by(username=form.username.data).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('user_management'))
        if User.query.filter_by(email=form.email.data).first():
            flash('Email already exists.', 'danger')
            return redirect(url_for('user_management'))
        
        user = User(
            username=form.username.data,
            email=form.email.data,
            password_hash=generate_password_hash(form.password.data),
            is_admin=form.is_admin.data
        )
        db.session.add(user)
        db.session.commit()
        flash('User created successfully!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{error}", 'danger')
    
    return redirect(url_for('user_management'))

@app.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    form = UserForm(obj=user)
    if form.validate_on_submit():
        # Check if username or email already exists but skip current user
        username_check = User.query.filter_by(username=form.username.data).first()
        if username_check and username_check.id != user_id:
            flash('Username already exists.', 'danger')
            return redirect(url_for('user_management'))
            
        email_check = User.query.filter_by(email=form.email.data).first()
        if email_check and email_check.id != user_id:
            flash('Email already exists.', 'danger')
            return redirect(url_for('user_management'))
        
        user.username = form.username.data
        user.email = form.email.data
        user.is_admin = form.is_admin.data
        
        # Only update password if a new one is provided
        if form.password.data:
            user.password_hash = generate_password_hash(form.password.data)
        
        db.session.commit()
        flash('User updated successfully!', 'success')
        return redirect(url_for('user_management'))
    
    return render_template('user_management.html', users=User.query.all(), form=form, edit_user=user)

@app.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Prevent deleting self
    if user_id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('user_management'))
    
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully!', 'success')
    return redirect(url_for('user_management'))

# Soldering Tip Requisition routes
@app.route('/soldering-tips')
@login_required
def soldering_tips():
    search_query = request.args.get('search', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = SolderingTip.query
    
    if search_query:
        query = query.filter(
            SolderingTip.machine_name.ilike(f'%{search_query}%')
        )
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(SolderingTip.date.between(start_date, end_date))
        except ValueError:
            flash('Invalid date format. Please use YYYY-MM-DD', 'danger')
    
    tips = query.order_by(SolderingTip.date.desc()).all()
    form = SolderingTipForm()
    
    # Get all usernames for the engineer_name dropdown
    form.engineer_name.choices = [(user.username, user.username) for user in User.query.all()]
    
    return render_template('soldering_tip.html', tips=tips, form=form, search_query=search_query, start_date=start_date, end_date=end_date)

@app.route('/soldering-tips/add', methods=['POST'])
@login_required
def add_soldering_tip():
    form = SolderingTipForm()
    
    # Get all usernames for the engineer_name dropdown
    form.engineer_name.choices = [(user.username, user.username) for user in User.query.all()]
    
    if form.validate_on_submit():
        new_tip = SolderingTip(
            machine_name=form.machine_name.data,
            engineer_name=form.engineer_name.data,
            personnel_name=form.personnel_name.data,
            shift=form.shift.data,
            date=form.date.data,
            created_at=datetime.utcnow()
        )
        db.session.add(new_tip)
        db.session.commit()
        flash('Soldering tip requisition added successfully!', 'success')
        return redirect(url_for('soldering_tips'))
    
    return render_template('soldering_tip.html', form=form)

@app.route('/soldering-tips/edit/<int:tip_id>', methods=['GET', 'POST'])
@login_required
def edit_soldering_tip(tip_id):
    tip = SolderingTip.query.get_or_404(tip_id)
    form = SolderingTipForm(obj=tip)
    
    # Get all usernames for the engineer_name dropdown
    form.engineer_name.choices = [(user.username, user.username) for user in User.query.all()]
    
    if form.validate_on_submit():
        tip.machine_name = form.machine_name.data
        tip.engineer_name = form.engineer_name.data
        tip.personnel_name = form.personnel_name.data
        tip.shift = form.shift.data
        tip.date = form.date.data
        db.session.commit()
        flash('Soldering tip requisition updated successfully!', 'success')
        return redirect(url_for('soldering_tips'))
    
    return render_template('soldering_tip.html', 
                           tips=SolderingTip.query.order_by(SolderingTip.date.desc()).all(), 
                           form=form, 
                           edit_tip=tip)

@app.route('/soldering-tips/delete/<int:tip_id>', methods=['POST'])
@login_required
def delete_soldering_tip(tip_id):
    tip = SolderingTip.query.get_or_404(tip_id)
    db.session.delete(tip)
    db.session.commit()
    flash('Soldering tip requisition deleted successfully!', 'success')
    return redirect(url_for('soldering_tips'))

# Machine Calibration Scheduler routes
@app.route('/machine-calibrations')
@login_required
def machine_calibrations():
    search_query = request.args.get('search', '')
    if search_query:
        calibrations = MachineCalibration.query.filter(
            MachineCalibration.machine_name.ilike(f'%{search_query}%')
        ).all()
    else:
        calibrations = MachineCalibration.query.all()
    form = MachineCalibrationForm()
    return render_template('machine_calibration.html', calibrations=calibrations, form=form, search_query=search_query)

@app.route('/machine-calibrations/add', methods=['GET', 'POST'])
@login_required
def add_machine_calibration():
    form = MachineCalibrationForm()
    if form.validate_on_submit():
        # Calculate the next calibration date based on the frequency
        if form.date.data:
            next_calibration_date = form.date.data + timedelta(days=form.days_per_calibration.data)
        else:
            next_calibration_date = datetime.now() + timedelta(days=form.days_per_calibration.data)
        
        calibration = MachineCalibration(
            machine_name=form.machine_name.data,
            days_per_calibration=form.days_per_calibration.data,
            location_line=form.location_line.data,
            operator_name=form.operator_name.data,
            date=next_calibration_date
        )
        db.session.add(calibration)
        db.session.commit()
        flash('Calibration schedule added successfully!', 'success')
        return redirect(url_for('machine_calibrations'))
    return render_template('machine_calibration.html', form=form, edit_calibration=None)

@app.route('/machine-calibrations/edit/<int:calibration_id>', methods=['GET', 'POST'])
@login_required
def edit_machine_calibration(calibration_id):
    calibration = MachineCalibration.query.get_or_404(calibration_id)
    form = MachineCalibrationForm(obj=calibration)
    
    if form.validate_on_submit():
        # Calculate the next calibration date based on the frequency
        if form.date.data:
            next_calibration_date = form.date.data + timedelta(days=form.days_per_calibration.data)
        else:
            next_calibration_date = datetime.now() + timedelta(days=form.days_per_calibration.data)
        
        calibration.machine_name = form.machine_name.data
        calibration.days_per_calibration = form.days_per_calibration.data
        calibration.location_line = form.location_line.data
        calibration.operator_name = form.operator_name.data
        calibration.date = next_calibration_date
        db.session.commit()
        flash('Calibration schedule updated successfully!', 'success')
        return redirect(url_for('machine_calibrations'))
    
    return render_template('machine_calibration.html', form=form, edit_calibration=calibration)

@app.route('/machine-calibrations/delete/<int:calibration_id>', methods=['POST'])
@login_required
def delete_machine_calibration(calibration_id):
    calibration = MachineCalibration.query.get_or_404(calibration_id)
    db.session.delete(calibration)
    db.session.commit()
    flash('Machine calibration schedule deleted successfully!', 'success')
    return redirect(url_for('machine_calibrations'))

@app.route('/overtime-logbook')
@login_required
def overtime_logbook():
    search_query = request.args.get('search', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    page = request.args.get('page', 1, type=int)
    form = OvertimeLogbookForm()
    form.employee_name.choices = [(user.id, user.username) for user in User.query.all()]
    form.employee_name.coerce = int

    # Get current month's logs for chart
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Calculate total hours per employee for the current month
    monthly_overtime = db.session.query(
        OvertimeLogbook.employee_name,
        db.func.sum(OvertimeLogbook.hours).label('total_hours')
    ).filter(
        db.extract('month', OvertimeLogbook.date) == current_month,
        db.extract('year', OvertimeLogbook.date) == current_year
    ).group_by(OvertimeLogbook.employee_name).all()
    
    # Convert to format for chart
    employee_names = [data[0] for data in monthly_overtime]
    total_hours = [float(data[1]) for data in monthly_overtime]
    
    query = OvertimeLogbook.query
    
    if search_query:
        query = query.filter(
            OvertimeLogbook.employee_name.ilike(f'%{search_query}%')
        )
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(OvertimeLogbook.date.between(start_date, end_date))
        except ValueError:
            flash('Invalid date format. Please use YYYY-MM-DD.', 'danger')
    
    logs = query.order_by(OvertimeLogbook.date.desc()).paginate(
        page=page,
        per_page=10,
        error_out=False
    )
    
    return render_template('overtime_logbook.html',
                         logs=logs,
                         form=form,
                         employee_names=json.dumps(employee_names),
                         total_hours=json.dumps(total_hours),
                         search_query=search_query,
                         start_date=start_date,
                         end_date=end_date)


@app.route('/overtime-logbook/add', methods=['POST'])
@login_required
def add_overtime_log():
    form = OvertimeLogbookForm()
    form.employee_name.choices = [(user.id, user.username) for user in User.query.all()]
    form.employee_name.coerce = int
    if form.validate_on_submit():
        log = OvertimeLogbook(
            employee_name=(User.query.get(form.employee_name.data).username if isinstance(form.employee_name.data, int) and User.query.get(form.employee_name.data) else form.employee_name.data),
            date=form.date.data,
            hours=form.hours.data
        )
        db.session.add(log)
        db.session.commit()
        flash('Overtime log added successfully!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{error}", 'danger')
    
    return redirect(url_for('overtime_logbook'))

@app.route('/overtime-logbook/edit/<int:log_id>', methods=['GET', 'POST'])
@login_required
def edit_overtime_log(log_id):
    log = OvertimeLogbook.query.get_or_404(log_id)
    form = OvertimeLogbookForm(obj=log)
    form.employee_name.choices = [(user.id, user.username) for user in User.query.all()]
    form.employee_name.coerce = int
    if form.validate_on_submit():
        log.employee_name = (User.query.get(form.employee_name.data).username if isinstance(form.employee_name.data, int) and User.query.get(form.employee_name.data) else form.employee_name.data)
        log.date = form.date.data
        log.hours = form.hours.data
        
        db.session.commit()
        flash('Overtime log updated successfully!', 'success')
        return redirect(url_for('overtime_logbook'))
    
    return render_template('overtime_logbook.html', 
                           logs=OvertimeLogbook.query.order_by(OvertimeLogbook.date.desc()).all(), 
                           form=form, 
                           edit_log=log)

@app.route('/overtime-logbook/delete/<int:log_id>', methods=['POST'])
@login_required
def delete_overtime_log(log_id):
    log = OvertimeLogbook.query.get_or_404(log_id)
    db.session.delete(log)
    db.session.commit()
    flash('Overtime log deleted successfully!', 'success')
    return redirect(url_for('overtime_logbook'))

# Equipment Downtime routes
@app.route('/equipment-downtime')
@login_required
def equipment_downtime():
    search_query = request.args.get('search', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = EquipmentDowntime.query
    
    if search_query:
        query = query.filter(
            EquipmentDowntime.equipment_name.ilike(f'%{search_query}%')
        )
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(EquipmentDowntime.date.between(start_date, end_date))
        except ValueError:
            flash('Invalid date format. Please use YYYY-MM-DD', 'danger')
    
    downtimes = query.order_by(EquipmentDowntime.date.desc()).all()
    form = EquipmentDowntimeForm()
    return render_template('equipment_downtime.html', downtimes=downtimes, form=form, search_query=search_query, start_date=start_date, end_date=end_date)

@app.route('/equipment-downtime/add', methods=['POST'])
@login_required
def add_equipment_downtime():
    form = EquipmentDowntimeForm()
    if form.validate_on_submit():
        downtime = EquipmentDowntime(
            equipment_name=form.equipment_name.data,
            product_name=form.product_name.data,
            issue=form.issue.data,
            downtime_minutes=form.downtime_minutes.data,
            shift=form.shift.data,
            action_taken=form.action_taken.data,
            date=form.date.data
        )
        db.session.add(downtime)
        db.session.commit()
        flash('Equipment downtime record added successfully!', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{error}", 'danger')
    
    return redirect(url_for('equipment_downtime'))

@app.route('/equipment-downtime/edit/<int:downtime_id>', methods=['GET', 'POST'])
@login_required
def edit_equipment_downtime(downtime_id):
    downtime = EquipmentDowntime.query.get_or_404(downtime_id)
    form = EquipmentDowntimeForm(obj=downtime)
    
    if form.validate_on_submit():
        downtime.equipment_name = form.equipment_name.data
        downtime.product_name = form.product_name.data
        downtime.issue = form.issue.data
        downtime.downtime_minutes = form.downtime_minutes.data
        downtime.shift = form.shift.data
        downtime.action_taken = form.action_taken.data
        downtime.date = form.date.data
        
        db.session.commit()
        flash('Equipment downtime record updated successfully!', 'success')
        return redirect(url_for('equipment_downtime'))
    
    return render_template('equipment_downtime.html', 
                           downtimes=EquipmentDowntime.query.order_by(EquipmentDowntime.date.desc()).all(), 
                           form=form, 
                           edit_downtime=downtime)

@app.route('/equipment-downtime/delete/<int:downtime_id>', methods=['POST'])
@login_required
def delete_equipment_downtime(downtime_id):
    downtime = EquipmentDowntime.query.get_or_404(downtime_id)
    db.session.delete(downtime)
    db.session.commit()
    flash('Equipment downtime record deleted successfully!', 'success')
    return redirect(url_for('equipment_downtime'))

# Maintenance and Abnormality Report routes
@app.route('/maintenance_report')
@login_required
def maintenance_report():
    return render_template('maintenance_report.html')

# Reports routes
@app.route('/reports')
@login_required
def reports():
    form = ReportForm()
    return render_template('reports.html', form=form)

@app.route('/reports/generate', methods=['GET', 'POST'])
@login_required
def generate_report():
    form = ReportForm()
    
    if form.validate_on_submit():
        report_type = form.report_type.data
        start_date = form.start_date.data
        end_date = form.end_date.data
        
        if report_type == 'soldering_tips':
            records = SolderingTip.query.filter(SolderingTip.date.between(start_date, end_date)).all()
            headers = ['Machine Name', 'Engineer Name', 'Personnel Name', 'Shift', 'Date']
        
        elif report_type == 'machine_calibrations':
            records = MachineCalibration.query.filter(MachineCalibration.created_at.between(start_date, end_date)).all()
            headers = ['Machine Name', 'Days per Calibration', 'Location/Line', 'Operator Name', 'Date']
        
        elif report_type == 'overtime_logbook':
            records = OvertimeLogbook.query.filter(OvertimeLogbook.date.between(start_date, end_date)).all()
            headers = ['Employee Name', 'Date', 'Hours', 'Created At']
        
        elif report_type == 'equipment_downtime':
            records = EquipmentDowntime.query.filter(EquipmentDowntime.date.between(start_date, end_date)).all()
            headers = ['Equipment Name', 'Product Name', 'Issue', 'Downtime (min)', 'Shift', 'Action Taken', 'Date', 'Created At']
            
        elif report_type == 'maintenance_reports':
            records = MaintenanceReport.query.filter(MaintenanceReport.created_at.between(start_date, end_date)).all()
            headers = ['Model ID', 'Model Name', 'Client Name', 'Station', 'Affected Component', 'Quantity', 'Problem Description', 'Status', 'Created At']
        
        # Return the data for preview
        return render_template('reports.html', 
            form=form,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            records=records,
            headers=headers,
            preview=True)

    return render_template('reports.html', form=form)

@app.route('/reports/download', methods=['POST'])
@login_required
def download_report():
    report_type = request.form.get('report_type')
    start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
    end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
    
    # Ensure end date is end of the day
    end_date = datetime.combine(end_date, datetime.max.time())
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Set up headers and styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    
    if report_type == 'soldering_tips':
        records = SolderingTip.query.filter(SolderingTip.date.between(start_date, end_date)).all()
        headers = ['Machine Name', 'Engineer Name', 'Personnel Name', 'Shift', 'Date', 'Created At']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.machine_name)
            ws.cell(row=row, column=2, value=record.engineer_name)
            ws.cell(row=row, column=3, value=record.personnel_name)
            ws.cell(row=row, column=4, value=record.shift)
            ws.cell(row=row, column=5, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=6, value=record.created_at.strftime('%Y-%m-%d %H:%M'))
        
        filename = f'soldering_tips_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
        
    elif report_type == 'machine_calibrations':
        records = MachineCalibration.query.filter(MachineCalibration.created_at.between(start_date, end_date)).all()
        headers = ['Machine Name', 'Calibration Frequency', 'Location/Line', 'Operator Name', 'Created At']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.machine_name)
            ws.cell(row=row, column=2, value=f'Every {record.days_per_calibration} days')
            ws.cell(row=row, column=3, value=record.location_line)
            ws.cell(row=row, column=4, value=record.operator_name)
            ws.cell(row=row, column=5, value=record.created_at.strftime('%Y-%m-%d %H:%M'))
        
        filename = f'machine_calibrations_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
    elif report_type == 'overtime_logbook':
        records = OvertimeLogbook.query.filter(OvertimeLogbook.date.between(start_date, end_date)).all()
        headers = ['Employee Name', 'Date', 'Hours', 'Created At']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.employee_name)
            ws.cell(row=row, column=2, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=3, value=record.hours)
            ws.cell(row=row, column=4, value=record.created_at.strftime('%Y-%m-%d %H:%M'))
        
        filename = f'overtime_logbook_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
        
    elif report_type == 'equipment_downtime':
        records = EquipmentDowntime.query.filter(EquipmentDowntime.date.between(start_date, end_date)).all()
        headers = ['Equipment Name', 'Product Name', 'Issue', 'Downtime (min)', 'Shift', 'Action Taken', 'Date', 'Created At']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.equipment_name)
            ws.cell(row=row, column=2, value=record.product_name)
            ws.cell(row=row, column=3, value=record.issue)
            ws.cell(row=row, column=4, value=record.downtime_minutes)
            ws.cell(row=row, column=5, value=record.shift)
            ws.cell(row=row, column=6, value=record.action_taken)
            ws.cell(row=row, column=7, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=8, value=record.created_at.strftime('%Y-%m-%d %H:%M'))
        
        filename = f'equipment_downtime_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    
    elif report_type == 'maintenance_reports':
        records = MaintenanceReport.query.filter(MaintenanceReport.created_at.between(start_date, end_date)).all()
        headers = ['Model ID', 'Model Name', 'Client Name', 'Station', 'Affected Component', 'Quantity', 'Problem Description', 'Status', 'Created At', 'Evidence Image']
        from openpyxl.drawing.image import Image as XLImage
        import os
        from PIL import Image as PILImage
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        temp_image_paths = []  # To store temp files for later cleanup
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.model_id)
            ws.cell(row=row, column=2, value=record.model_name)
            ws.cell(row=row, column=3, value=record.client_name)
            ws.cell(row=row, column=4, value=record.station)
            ws.cell(row=row, column=5, value=record.affected_component)
            ws.cell(row=row, column=6, value=record.quantity)
            ws.cell(row=row, column=7, value=record.problem_description)
            ws.cell(row=row, column=8, value=record.status)
            ws.cell(row=row, column=9, value=record.created_at.strftime('%Y-%m-%d %H:%M'))

            # Insert evidence image if available
            evidence_path = record.evidence
            if evidence_path:
                image_full_path = os.path.join(app.config['UPLOAD_FOLDER'], evidence_path)
                if os.path.isfile(image_full_path):
                    try:
                        # Optionally resize for Excel
                        with PILImage.open(image_full_path) as img:
                            max_dim = 120
                            img.thumbnail((max_dim, max_dim))
                            temp_path = image_full_path + '_excel_tmp.png'
                            img.save(temp_path, format='PNG')
                        img_xl = XLImage(temp_path)
                        img_xl.width, img_xl.height = img.size
                        cell_ref = f'J{row}'
                        ws.add_image(img_xl, cell_ref)
                        temp_image_paths.append(temp_path)  # Defer cleanup
                    except Exception as e:
                        ws.cell(row=row, column=10, value='Image error')
                else:
                    ws.cell(row=row, column=10, value='Not found')
            else:
                ws.cell(row=row, column=10, value='No image')

        filename = f'maintenance_reports_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    
    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Clean up temp image files if any
    if 'temp_image_paths' in locals():
        for temp_path in temp_image_paths:
            try:
                os.remove(temp_path)
            except Exception:
                pass
    
    # Return the Excel file for download
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# API routes for dashboard charts
@app.route('/api/dashboard/soldering_tips_data')
@login_required
def soldering_tips_data():
    # Get data for past 30 days
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    # Query data
    data = SolderingTip.query.filter(SolderingTip.date.between(start_date, end_date)).all()
    
    # Format data for chart - count by day
    dates = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(31)]
    counts = [0] * 31
    
    for tip in data:
        day_index = (tip.date.date() - start_date.date()).days
        if 0 <= day_index < 31:
            counts[day_index] += 1
    
    return jsonify({
        'labels': dates,
        'datasets': [{
            'label': 'Soldering Tips',
            'data': counts,
            'borderColor': 'rgba(75, 192, 192, 1)',
            'backgroundColor': 'rgba(75, 192, 192, 0.2)',
            'fill': True
        }]
    })

@app.route('/api/dashboard/overtime_data')
@login_required
def overtime_data():
    # Get data for past 30 days
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    # Query data
    data = OvertimeLogbook.query.filter(OvertimeLogbook.date.between(start_date, end_date)).all()
    
    # Format data for chart - sum hours by day
    dates = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(31)]
    hours = [0] * 31
    
    for log in data:
        day_index = (log.date.date() - start_date.date()).days
        if 0 <= day_index < 31:
            hours[day_index] += log.hours
    
    return jsonify({
        'labels': dates,
        'datasets': [{
            'label': 'Overtime Hours',
            'data': hours,
            'borderColor': 'rgba(153, 102, 255, 1)',
            'backgroundColor': 'rgba(153, 102, 255, 0.2)',
            'fill': True
        }]
    })

@app.route('/api/dashboard/downtime_data')
@login_required
def downtime_data():
    # Get data for past 30 days
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    # Query data
    data = EquipmentDowntime.query.filter(EquipmentDowntime.date.between(start_date, end_date)).all()
    
    # Format data for chart - sum downtime by day
    dates = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(31)]
    downtimes = [0] * 31
    
    for record in data:
        day_index = (record.date.date() - start_date.date()).days
        if 0 <= day_index < 31:
            downtimes[day_index] += record.downtime_minutes
    
    return jsonify({
        'labels': dates,
        'datasets': [{
            'label': 'Equipment Downtime (minutes)',
            'data': downtimes,
            'borderColor': 'rgba(255, 99, 132, 1)',
            'backgroundColor': 'rgba(255, 99, 132, 0.2)',
            'fill': True
        }]
    })

@app.route('/api/dashboard/calibration_data')
@login_required
def calibration_data():
    # Get all calibration records
    calibrations = MachineCalibration.query.all()
    
    # Group by days per calibration
    days_data = {}
    for cal in calibrations:
        days = cal.days_per_calibration
        if days in days_data:
            days_data[days] += 1
        else:
            days_data[days] = 1
    
    # Sort by days
    sorted_days = sorted(days_data.keys())
    
    return jsonify({
        'labels': [f'{d} days' for d in sorted_days],
        'datasets': [{
            'label': 'Machines by Calibration Frequency',
            'data': [days_data[d] for d in sorted_days],
            'backgroundColor': [
                'rgba(255, 99, 132, 0.6)',
                'rgba(54, 162, 235, 0.6)',
                'rgba(255, 206, 86, 0.6)',
                'rgba(75, 192, 192, 0.6)',
                'rgba(153, 102, 255, 0.6)',
                'rgba(255, 159, 64, 0.6)'
            ],
            'borderColor': [
                'rgba(255, 99, 132, 1)',
                'rgba(54, 162, 235, 1)',
                'rgba(255, 206, 86, 1)',
                'rgba(75, 192, 192, 1)',
                'rgba(153, 102, 255, 1)',
                'rgba(255, 159, 64, 1)'
            ],
            'borderWidth': 1
        }]
    })

@app.route('/api/monthly-overtime-data')
@login_required
def monthly_overtime_data():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    period = request.args.get('period', 'month')  # Default to month
    
    today = datetime.now()
    
    if not from_date or not to_date:
        if period == '15days':
            from_date = (today - timedelta(days=15)).strftime('%Y-%m-%d')
            to_date = today.strftime('%Y-%m-%d')
        else:  # Default to month
            from_date = datetime(today.year, today.month, 1).strftime('%Y-%m-%d')
            to_date = (datetime(today.year, today.month, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            to_date = to_date.strftime('%Y-%m-%d')
    
    # Get monthly overtime data
    monthly_overtime = db.session.query(
        OvertimeLogbook.employee_name,
        db.func.sum(OvertimeLogbook.hours).label('total_hours')
    ).filter(
        OvertimeLogbook.date.between(from_date, to_date)
    ).group_by(
        OvertimeLogbook.employee_name
    ).order_by(
        OvertimeLogbook.employee_name
    ).all()
    
    # Prepare data for chart
    labels = [record.employee_name for record in monthly_overtime]
    data = [float(record.total_hours) for record in monthly_overtime]
    
    return jsonify({
        'labels': labels,
        'data': data
    })

@app.route('/maintenance-reports', methods=['GET', 'POST'])
@login_required
def maintenance_reports():
    edit_id = request.args.get('edit_id', type=int)
    report_to_edit = None
    if edit_id:
        report_to_edit = MaintenanceReport.query.get(edit_id)
        form = MaintenanceReportForm(obj=report_to_edit)
    else:
        form = MaintenanceReportForm()

    if form.validate_on_submit():
        evidence_file = request.files.get('evidence')
        filename = None
        if evidence_file and evidence_file.filename and allowed_file(evidence_file.filename):
            filename = secure_filename(evidence_file.filename)
            evidence_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        if edit_id and report_to_edit:
            # Update existing report
            form.populate_obj(report_to_edit)
            if filename:
                report_to_edit.evidence = filename  # Only assign the filename if a new file was uploaded!
            # Do NOT overwrite evidence if no file is uploaded!
            db.session.commit()
            flash('Maintenance report updated successfully!', 'success')
        else:
            # Insert new report
            new_report = MaintenanceReport()
            form.populate_obj(new_report)
            # Set client_id based on selected client_name (assume value is client_id)
            new_report.client_id = form.client_name.data
            if filename:
                new_report.evidence = filename
            db.session.add(new_report)
            db.session.commit()
            flash('Maintenance report added successfully!', 'success')
        return redirect(url_for('maintenance_reports'))

    # Fetch reports for the list
    page = request.args.get('page', 1, type=int)
    reports = MaintenanceReport.query.order_by(MaintenanceReport.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    return render_template('maintenance_reports.html', form=form, reports=reports, edit_id=edit_id)
@app.route('/maintenance-reports/<int:report_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_maintenance_report(report_id):
    report = MaintenanceReport.query.get_or_404(report_id)
    form = MaintenanceReportForm(obj=report)
    
    if form.validate_on_submit():
        if edit_id and report_to_edit:
            # Update existing report
            form.populate_obj(report_to_edit)
            # Handle evidence upload
            evidence_file = request.files.get('evidence')
            if evidence_file and evidence_file.filename and allowed_file(evidence_file.filename):
                filename = secure_filename(evidence_file.filename)
                evidence_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                report_to_edit.evidence = filename  # Only assign the filename!
            # Do NOT overwrite evidence if no file is uploaded!
            db.session.commit()
            flash('Maintenance report updated successfully!', 'success')
        else:
            # Insert new report
            new_report = MaintenanceReport()
            form.populate_obj(new_report)
            evidence_file = request.files.get('evidence')
            if evidence_file and evidence_file.filename and allowed_file(evidence_file.filename):
                filename = secure_filename(evidence_file.filename)
                evidence_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                new_report.evidence = filename
            db.session.add(new_report)
            db.session.commit()
            flash('Maintenance report added successfully!', 'success')
        return redirect(url_for('maintenance_reports'))

@app.route('/maintenance-reports/<int:report_id>/delete', methods=['POST'])
@login_required
def delete_maintenance_report(report_id):
    report = MaintenanceReport.query.get_or_404(report_id)
    if report.evidence:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], report.evidence))
        except:
            pass
    db.session.delete(report)
    db.session.commit()
    flash('Maintenance report deleted successfully!', 'success')
    return redirect(url_for('maintenance_reports'))



# Helper function for file uploads
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'jpg', 'jpeg', 'png', 'gif'}

@app.route('/api/check-calibrations')
@login_required
def check_calibrations():
    today = datetime.now()
    threshold_days = 7  # Notify if calibration is within 7 days
    
    upcoming_calibrations = MachineCalibration.query.filter(
        MachineCalibration.date.isnot(None),
        MachineCalibration.date <= today + timedelta(days=threshold_days)
    ).all()
    
    return jsonify({
        'upcoming_calibrations': [
            {
                'id': cal.id,
                'machine_name': cal.machine_name,
                'date': cal.date.strftime('%Y-%m-%d'),
                'days_until': (cal.date - today).days
            }
            for cal in upcoming_calibrations
        ]
    })


@app.route('/it-inventory', methods=['GET', 'POST'])
@login_required
def it_inventory():
    edit_id = request.args.get('edit_id')
    form = ITInventoryForm()

    if edit_id:
        item_to_edit = ITInventory.query.get(edit_id)
        if item_to_edit and request.method == 'GET':
            # Prefill form with item data
            form.item_name.data = item_to_edit.item_name
            form.total_quantity.data = item_to_edit.total_quantity
            form.acquired_qty.data = item_to_edit.acquired_qty

    if form.validate_on_submit():
        if edit_id:
            # Update existing item
            item_to_edit = ITInventory.query.get(edit_id)
            if item_to_edit:
                item_to_edit.item_name = form.item_name.data
                item_to_edit.total_quantity = form.total_quantity.data
                item_to_edit.acquired_qty = form.acquired_qty.data
                db.session.commit()
                flash('Item updated successfully!', 'success')
        else:
            # Insert new item
            new_item = ITInventory(
                item_name=form.item_name.data,
                total_quantity=form.total_quantity.data,
                acquired_qty=form.acquired_qty.data
            )
            db.session.add(new_item)
            db.session.commit()
            flash('Item added successfully!', 'success')
        return redirect(url_for('it_inventory'))

    inventory_items = ITInventory.query.all()
    return render_template('it_inventory.html', form=form, inventory_items=inventory_items, edit_id=edit_id)

@app.route('/it-inventory/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
def edit_it_inventory(item_id):
    item = ITInventory.query.get_or_404(item_id)
    form = ITInventoryForm(obj=item)
    if form.validate_on_submit():
        item.item_name = form.item_name.data
        item.total_quantity = form.total_quantity.data
        item.acquired_qty = form.acquired_qty.data
        db.session.commit()
        flash('Item updated successfully!', 'success')
        return redirect(url_for('it_inventory'))
    return render_template('edit_it_inventory.html', form=form, item=item)

@app.route('/it-inventory/delete/<int:item_id>', methods=['POST', 'GET'])
@login_required
def delete_it_inventory(item_id):
    item = ITInventory.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('Item deleted successfully!', 'success')
    return redirect(url_for('it_inventory'))